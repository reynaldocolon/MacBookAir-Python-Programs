import openpyxl
import warnings
import re
import calendar
import os
import tkinter as tk
from tkinter import filedialog
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

warnings.filterwarnings("ignore", category=UserWarning)
def select_file():
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    file_path = filedialog.askopenfilename(title="Select WIC Report", filetypes=[("Excel files", "*.xlsx")])
    root.destroy()
    return file_path

def get_report_data(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb["Main"] if "Main" in wb.sheetnames else wb.active
    
    # 1. Full Account Extraction (Captures "test" suffix)
    raw_a5 = str(sheet["A5"].value or "").strip()
    g_match = re.search(r'(G-.*)', raw_a5, re.IGNORECASE)
    g_number = g_match.group(1).strip() if g_match else "G-Number Missing"
    
    # 2. Date Ranges
    filename = os.path.basename(file_path).upper()
    month_map = {'JAN':'01','FEB':'02','MAR':'03','APR':'04','MAY':'05','JUN':'06',
                 'JUL':'07','AUG':'08','SEP':'09','OCT':'10','NOV':'11','DEC':'12'}
    month_num = next((m_num for k, m_num in month_map.items() if k in filename), "01")
    year = re.search(r'20\d{2}', filename).group(0) if re.search(r'20\d{2}', filename) else "2026"
    last_day = calendar.monthrange(int(year), int(month_num))[1]
    dates = {"from": f"{month_num}/01/{year}", "through": f"{month_num}/{last_day}/{year}"}
    def val(r, c): return int(float(sheet.cell(r, c).value or 0) + 0.5)

    # 3. 35-Field Sequence Mapping
    data_sequence = [
        val(16,6), val(16,4), val(16,3), val(16,5), val(16,8), # Salaries
        val(37,6), val(37,4), val(37,3), val(37,5), val(37,8), # Fringe
        0,                                                    # 11: Subtotal
        0, 0,                                                 # 12-13: Construction/Equip
        val(96,3),                                            # 14: Facility (14,724)
        0, 0,                                                 # 15-16: Prof/Subaward
        val(52,6), val(52,4), val(52,3), val(52,5),           # 17-20: Supplies (1,917 in CS)
        val(69,6), val(69,3), val(69,5),                      # 21-23: Travel (53 in NE)
        val(77,6),                                            # 24: Training
        val(88,6), val(88,7), val(88,4), val(88,3), val(88,8),# 25-29: Other
        0, 0, 0, 0, 0, 0                                      # 30-35: Padding
    ]
    
    # Calculate Total for Summary Display
    grand_total = sum(data_sequence)
            
    return g_number, data_sequence, dates, grand_total
def run_automation(g_number, data, periods, total):
    options = Options()
    options.add_experimental_option("detach", True)
    driver = webdriver.Chrome(options=options)
    
    driver.get("https://dohsage.intelligrants.com/IGXLogin")
    
    # Restore the visual summary you requested
    print("="*60)
    print("📄 DATA READY FOR INJECTION:")
    print(f"   Account: {g_number}")
    print(f"   Dates:   {periods['from']} - {periods['through']}")
    print(f"   Total:   ${total:,}")
    print("="*60)
    
    input("\n👉 Log in, handle Account manually if needed, THEN press ENTER...")

    iframes = driver.find_elements(By.TAG_NAME, "iframe")
    # STEP 1: Injection of Dates (and Account attempt)
    for frame_index in range(-1, len(iframes)):
        if frame_index == -1: driver.switch_to.default_content()
        else: driver.switch_to.frame(frame_index)
        
        # Injection for Dates
        for xpath, val_str in [("//input[contains(@id, 'From')]", periods['from']), ("//input[contains(@id, 'To')]", periods['through'])]:
            try:
                el = driver.find_element(By.XPATH, xpath)
                driver.execute_script("arguments[0].removeAttribute('readonly');", el)
                el.clear()
                el.send_keys(val_str)
                driver.execute_script("arguments[0].dispatchEvent(new Event('change'));", el)
            except: continue

    # STEP 2: Injection of 35 Financial Fields
    search_query = "//input[contains(@aria-label, 'monPrdGrant') or contains(@data-igx-alias, 'monPrdGrant')]"
    
    for frame_index in range(-1, len(iframes)):
        if frame_index == -1: driver.switch_to.default_content()
        else: driver.switch_to.frame(frame_index)
        found_fields = driver.find_elements(By.XPATH, search_query)
        
        if found_fields:
            print(f"🚀 Filling {len(data)} fields...") # Now correctly shows 35
            for i, value in enumerate(data):
                if i < len(found_fields):
                    driver.execute_script("""
                        arguments[0].removeAttribute('disabled');
                        arguments[0].removeAttribute('readonly');
                        arguments[0].value = arguments[1];
                        arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                    """, found_fields[i], str(value))
            break
if __name__ == "__main__":
    while True:
        try:
            path = select_file()
            if not path: break
            g_num, values, dates, total_sum = get_report_data(path)
            run_automation(g_num, values, dates, total_sum)
            if input("\n🔄 Run another? (y/n): ").lower() != 'y': break
        except Exception as e:
            print(f"⚠️ Error: {e}")
            break
